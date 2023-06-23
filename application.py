import sys
import openpyxl
import multiprocessing
from PyQt5.QtWidgets import QApplication
from view.main_view import MainWindow
from model.code_manipulator import CodeManipulator
from model.variable_combiner import Combiner


class Utility:
    def scan_variables():
        try:
            CodeManipulator.codeAddress = Application.codeAddress
            variables_list = CodeManipulator.scan_variables()
            if variables_list:
                Application.view.clear_table()
            for this in variables_list:
                name, var = this
                Application.view.append_new_row(header=name, variable=var)

        except FileNotFoundError:
            print(
                f"file not found at this address:\n{CodeManipulator.codeAddress}")

    def generate_inputs():
        table_obj = Application.view.tableWidget
        avail_vars = Combiner.find_avaliable_variables(table_obj)

        excel_fileName = Application.excelAddress
        avail_values = Combiner.load_variables_values(
            avail_vars, excel_fileName)
        result = Combiner.combine(avail_values)
        preview = "\n".join(list(map(lambda x: f"input{x}", result)))
        Application.view.plainTextEdit.setPlainText(preview)
        Application.generated_inputs = result
        Application.generated_inputs_labels = list(avail_values.keys())
        Application.view.pushButton_export.setEnabled(True)

    def export_generated_inputs(fileName):
        if not (Application.generate_inputs):
            return -1
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.cell(row=1, column=1, value="CaseStudyName")
        for i, label in enumerate(Application.generated_inputs_labels, start=2):
            worksheet.cell(row=1, column=i, value=label)

        for i, value_list in enumerate(Application.generated_inputs, start=2):
            worksheet.cell(row=i, column=1, value="Case"+str(i-1).zfill(8))
            for j, value in enumerate(value_list, start=2):
                worksheet.cell(row=i, column=j, value=value)
        workbook.save(fileName)

    def get_number_of_cores():
        return multiprocessing.cpu_count()

    def load_final_cases():
        result = []
        fileName = Application.excel_inputs_address
        if not fileName:
            return []
        workbook = openpyxl.load_workbook(fileName)
        worksheet = workbook.active
        labels = list(map(lambda x: x.value, worksheet[1]))
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            this_case = {}
            # print(row)
            for i, this_label in enumerate(labels):
                this_case[this_label] = row[i]
            result.append(this_case)
            # print(this_case)
        return result

    def generate_launcher():
        cores = Application.view.comboBox_cores.currentText()
        fileName = Application.output_folder + "/" + "Launcher.py"
        CodeManipulator.generate_launcher(fileName, cores)

    def generate_launcher_procedure():
        cases = Application.load_final_cases()
        CodeManipulator.codeAddress = Application.codeAddress
        for this_case in cases:
            # print(this_case)
            name = Application.output_folder + '/' + this_case["CaseStudyName"]
            del this_case["CaseStudyName"]
            data = this_case
            CodeManipulator.generate_case(name, data)
            # print(f"{name=}\n{data=}\n-----------")
        Application.generate_launcher()


class Application(Utility):
    qapp = None
    view = None
    output_folder = "."
    excel_inputs_address = None
    generated_inputs = None
    generated_inputs_labels = None

    @classmethod
    def init_ui(cls):
        cls.qapp = QApplication(sys.argv)
        cls.view = MainWindow(cls)

    def normal_run():
        Application.init_ui()
        Application.view.show()
        sys.exit(Application.qapp.exec())
